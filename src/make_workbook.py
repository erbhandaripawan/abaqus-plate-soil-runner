import argparse
import csv
import os

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, ScatterChart, Series
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config_io import case_name, load_config


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def read_csv_rows(path):
    if not os.path.exists(path):
        return []
    with open(path, "r", newline="") as f:
        return list(csv.reader(f))


def coerce_cell(value):
    if value is None or value == "":
        return value
    try:
        if isinstance(value, str) and value.strip() == "":
            return value
        return float(value)
    except (TypeError, ValueError):
        return value


def write_sheet(wb, title, rows):
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["No data found"])
        return ws
    ws.append(rows[0])
    for row in rows[1:]:
        ws.append([coerce_cell(value) for value in row])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1):
        max_len = 10
        for row in range(1, min(ws.max_row, 200) + 1):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                max_len = max(max_len, min(28, len(str(value)) + 2))
        ws.column_dimensions[get_column_letter(col)].width = max_len
    return ws


def add_decay_chart(wb, decay_ws):
    chart_ws = wb.create_sheet("Charts")
    chart_ws["A1"] = "Decay function: vertical deflection / surface deflection"
    chart_ws["A1"].font = Font(bold=True, size=14)
    chart_ws["A3"] = "Charts are native Excel objects and remain editable."

    if decay_ws.max_row < 3:
        return chart_ws

    chart = ScatterChart()
    chart.title = "Soil Deflection Decay"
    chart.x_axis.title = "Depth below soil surface"
    chart.y_axis.title = "U2(depth) / U2(surface)"
    chart.legend.position = "r"
    chart.style = 13

    headers = [cell.value for cell in decay_ws[1]]
    col_factor = headers.index("radius_factor") + 1
    col_depth = headers.index("depth") + 1
    col_ratio = headers.index("decay_ratio") + 1

    factors = []
    for row in range(2, decay_ws.max_row + 1):
        value = decay_ws.cell(row=row, column=col_factor).value
        if value not in factors:
            factors.append(value)

    helper = wb.create_sheet("DecayChartData")
    helper.sheet_state = "hidden"
    helper["A1"] = "depth"
    max_len = 0
    for idx, factor in enumerate(factors, start=2):
        helper.cell(row=1, column=idx).value = "{} rp".format(factor)
    depth_rows = {}
    for row in range(2, decay_ws.max_row + 1):
        depth = decay_ws.cell(row=row, column=col_depth).value
        if depth == "" or depth is None:
            continue
        if depth not in depth_rows:
            depth_rows[depth] = len(depth_rows) + 2
            helper.cell(row=depth_rows[depth], column=1).value = float(depth)
    for row in range(2, decay_ws.max_row + 1):
        factor = decay_ws.cell(row=row, column=col_factor).value
        depth = decay_ws.cell(row=row, column=col_depth).value
        ratio = decay_ws.cell(row=row, column=col_ratio).value
        if depth in depth_rows and factor in factors and ratio not in ("", None):
            helper.cell(row=depth_rows[depth], column=factors.index(factor) + 2).value = float(ratio)
            max_len = max(max_len, depth_rows[depth])

    xvalues = Reference(helper, min_col=1, min_row=2, max_row=max_len)
    for idx, factor in enumerate(factors, start=2):
        yvalues = Reference(helper, min_col=idx, min_row=2, max_row=max_len)
        series = Series(yvalues, xvalues, title_from_data=False, title="{} rp".format(factor))
        chart.series.append(series)
    chart_ws.add_chart(chart, "A5")
    return chart_ws


def add_line_chart(ws, title, x_col, y_cols, anchor, x_title, y_title):
    if ws.max_row < 3:
        return
    chart = LineChart()
    chart.title = title
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    cats = Reference(ws, min_col=x_col, min_row=2, max_row=ws.max_row)
    for col in y_cols:
        data = Reference(ws, min_col=col, min_row=1, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 9
    chart.width = 16
    ws.add_chart(chart, anchor)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--case", required=True)
    args = parser.parse_args()

    cfg = load_config(args.case)
    out_dir = cfg["output"]["directory"]
    os.makedirs(out_dir, exist_ok=True)
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Case", case_name(cfg)])
    summary.append(["Config", os.path.abspath(args.case)])
    summary.append(["Output directory", out_dir])
    summary.append(["Plate radius rp", cfg["geometry"]["plate_radius"]])
    summary.append(["Plate thickness", cfg["geometry"]["plate_thickness"]])
    summary.append(["Soil depth H_total", cfg["geometry"]["soil_depth"]])
    summary.append(["Boundary: bottom", "U2 = 0 at H_total"])
    summary.append(["Boundary: far radius", "U1 = 0, U2 = 0 at 100rp"])
    summary.append(["Boundary: axis", "U1 = 0 at r=0"])
    summary["A1"].font = Font(bold=True)
    summary.column_dimensions["A"].width = 26
    summary.column_dimensions["B"].width = 80

    disp_ws = write_sheet(wb, "Displacements", read_csv_rows(os.path.join(out_dir, "displacements.csv")))
    decay_ws = write_sheet(wb, "Decay", read_csv_rows(os.path.join(out_dir, "decay.csv")))
    stress_ws = write_sheet(wb, "Stresses", read_csv_rows(os.path.join(out_dir, "stresses.csv")))
    strain_ws = write_sheet(wb, "Strains", read_csv_rows(os.path.join(out_dir, "strains.csv")))
    result_ws = write_sheet(wb, "Plate Resultants", read_csv_rows(os.path.join(out_dir, "plate_resultants.csv")))

    add_decay_chart(wb, decay_ws)
    add_line_chart(result_ws, "Plate moments and shear resultants", 2, [3, 4, 5], "H2", "Radius", "Resultant")

    xlsx_path = os.path.join(out_dir, "{}_postprocessing.xlsx".format(case_name(cfg)))
    wb.save(xlsx_path)
    print("Workbook written: {}".format(xlsx_path))


if __name__ == "__main__":
    main()
